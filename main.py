import pandas as pd
import openpyxl as op
from openpyxl.styles import PatternFill

read_csv = pd.read_csv('excel_files/Levelling.csv')
read_csv.to_excel('excel_files/level.xlsx')


work_book = op.load_workbook('excel_files/level.xlsx')
ws = work_book.active

list_of_all_sights = []
list_of_Rise_n_Fall = []
count = 0


def calTotal(sheet, list_Len, columnPos):
    global count
    count = 0
    for i in range(2, 3+list_Len):
        col_cells = sheet.cell(row=i, column=columnPos).value
        if col_cells is not None:
            count += col_cells

    sheet.cell(row=i+2, column=columnPos).fill = PatternFill(bgColor='71FF33')
    sheet.cell(row=i+2, column=columnPos).value = count
    return count


# * Looping through the rows and storing all none empty cell in the list
for rows in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    individual_sights = [cell.value for cell in rows if cell.value]
    list_of_all_sights.append(individual_sights)

# print(list_of_all_sights)
# * Looping through the list_of_all_sight and calculating the Rise and Fall
for i, items in enumerate(list_of_all_sights):
    # * checking if the index exceeds the length of the list
    if i < (len(list_of_all_sights)-1):
        bs = list_of_all_sights[i][0]
        fs = list_of_all_sights[i+1][0]
        if len(list_of_all_sights[i]) == 2:
            bs = list_of_all_sights[i][0]
        if len(list_of_all_sights[i+1]) == 2:
            fs = list_of_all_sights[i+1][1]
        Rise_or_Fall = bs-fs
        list_of_Rise_n_Fall.append(Rise_or_Fall)

ws.insert_cols(5)
ws.insert_cols(6)
ws.delete_cols(1)
# print(list_of_Rise_n_Fall)

work_book.save('results.xlsx')

work_book1 = op.load_workbook('results.xlsx')
ws1 = work_book1.active


BM = eval(input('Enter Your BenchMark: '))

# * looping through the Rise/Fall column and adding the values in the list_of_Rise_n_Fall into their appropriate cell
ws1.cell(row=1, column=4).value = 'Rise/Fall'
for cols in ws1.iter_cols(min_col=4, min_row=3, max_col=4):
    for i, cells in enumerate(cols):
        cells.value = list_of_Rise_n_Fall[i]

# * calculating the Reduced Level by adding the previous value from the RL column to the corresponding value in the Rise/Fall column
ws1.cell(row=1, column=5).value = 'RL'
for i in range(2, 3+len(list_of_Rise_n_Fall)):
    if i == 2:
        ws1.cell(row=i, column=5).value = BM
    else:
        results = float(ws1.cell(
            row=i-1, column=5).value) + float(ws1.cell(
                row=i, column=4).value)

        ws1.cell(row=i, column=5).value = results
        # print(
        #     f'index={i}----rl={float(ws1.cell(row=i-1, column=5).value)} - rf={float(ws1.cell(row=i, column=4).value)} and results = {results} ')
# * Sum Total of the Back Sight(B.S), Fore Sight(F.S) and Reduce Level (R.L)
tot_BS = calTotal(ws1, len(list_of_Rise_n_Fall), 1)
tot_FS = count = calTotal(ws1, len(list_of_Rise_n_Fall), 3)
tot_RL = count = calTotal(ws1, len(list_of_Rise_n_Fall), 4)
# print()
# print(
#     f'total of BS = {tot_BS} \n total of FS = {tot_FS} \n total of RL = {tot_RL}')
work_book1.save('results.xlsx')
